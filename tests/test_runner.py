import threading
import time
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pytest
from openpyxl import Workbook

from tender_tracker.config import load_settings
from tender_tracker.excel import build_sheet_name
from tender_tracker.logging_utils import build_logger
from tender_tracker.models import CompanyRecord, PaymentRecord, SearchResultItem
from tender_tracker.runner import TenderTrackerApp, _map_parallel
from tender_tracker.state import RunStateStore
from tender_tracker.storage import LocalStorage
from tender_tracker.tender_client import TenderClientError


def test_map_parallel_respects_concurrency_bound_and_returns_all_results():
    lock = threading.Lock()
    active = 0
    max_active = 0

    def fn(item: int) -> int:
        nonlocal active, max_active
        with lock:
            active += 1
            max_active = max(max_active, active)
        time.sleep(0.05)
        with lock:
            active -= 1
        return item * 2

    items = list(range(6))
    results = {item: result for item, result, error in _map_parallel(fn, items, concurrency=3)}

    assert results == {item: item * 2 for item in items}
    assert 2 <= max_active <= 3


def test_map_parallel_yields_tender_client_error_as_sentinel_without_aborting():
    def fn(item: str) -> str:
        if item == "bad":
            raise TenderClientError("boom")
        return f"ok-{item}"

    outcomes = {item: (result, error) for item, result, error in _map_parallel(fn, ["a", "bad", "b"], concurrency=2)}

    assert outcomes["a"] == ("ok-a", None)
    assert outcomes["b"] == ("ok-b", None)
    result, error = outcomes["bad"]
    assert result is None
    assert isinstance(error, TenderClientError)
    assert str(error) == "boom"


def test_map_parallel_propagates_unexpected_exception():
    def fn(item: int) -> int:
        if item == 2:
            raise ValueError("unexpected")
        return item

    with pytest.raises(ValueError, match="unexpected"):
        list(_map_parallel(fn, [1, 2, 3], concurrency=3))


def test_map_parallel_with_empty_items_yields_nothing():
    assert list(_map_parallel(lambda item: item, [], concurrency=4)) == []


def _search_item(app_id: str, company_id: str, reg_id: str | None = None) -> SearchResultItem:
    return SearchResultItem(
        app_id=app_id,
        company_id=company_id,
        company_name="Co",
        tender_registration_number=reg_id,
        announcement_date=None,
        row_text="",
        page_number=1,
        total_pages=1,
    )


class FakeClient:
    def __init__(self, results_by_company: dict[str, list[SearchResultItem]], failing_company_ids: frozenset[str] = frozenset()):
        self.results_by_company = results_by_company
        self.failing_company_ids = failing_company_ids
        self.calls: list[str] = []
        self._lock = threading.Lock()

    def search_company(self, company: CompanyRecord, status_id: int) -> list[SearchResultItem]:
        with self._lock:
            self.calls.append(company.company_id)
        if company.company_id in self.failing_company_ids:
            raise TenderClientError(f"boom-{company.company_id}")
        return self.results_by_company.get(company.company_id, [])

    def initialize(self) -> None:
        pass

    retry_count = 0

    def fetch_tender_details_concurrent(self, targets, *, on_success, on_failure) -> None:
        for item in targets:
            on_success(
                PaymentRecord(
                    company_id=item.company_id,
                    company_name=item.company_name,
                    app_id=item.app_id,
                    tender_registration_number=item.tender_registration_number,
                    raw_amount="100.00",
                    cleaned_amount=100.0,
                    raw_payment_date="01/01/2026",
                    parsed_payment_date=None,
                    payment_exists=True,
                )
            )


def _make_app(tmp_path: Path, client: FakeClient, concurrency: int = 3) -> tuple[TenderTrackerApp, RunStateStore]:
    settings = load_settings("config/settings.yaml")
    settings.scraper.company_search_concurrency = concurrency
    settings.scraper.contract_status_ids = [10]
    storage = LocalStorage(tmp_path)
    state_store = RunStateStore(settings, storage)
    logger = build_logger(tmp_path / "test.log", "INFO")
    app = TenderTrackerApp(
        settings=settings,
        storage=storage,
        state_store=state_store,
        client=client,
        logger=logger,
        work_root=tmp_path / "work",
    )
    return app, state_store


def test_collect_targets_dedupes_across_companies_and_records_all_as_processed(tmp_path: Path):
    companies = [
        CompanyRecord(company_id="1", company_name="A", overdue_days_raw="1"),
        CompanyRecord(company_id="2", company_name="B", overdue_days_raw="1"),
        CompanyRecord(company_id="3", company_name="C", overdue_days_raw="1"),
    ]
    results_by_company = {
        "1": [_search_item("app-1", "1", "REG-1")],
        "2": [_search_item("app-1", "2", "REG-1"), _search_item("app-2", "2", "REG-2")],
        "3": [_search_item("app-3", "3")],
    }
    client = FakeClient(results_by_company)
    app, state_store = _make_app(tmp_path, client)
    state = state_store.create("run", {}, companies)

    targets = app._collect_targets(state, companies)

    assert sorted(item.app_id for item in targets) == ["app-1", "app-2", "app-3"]
    assert sorted(client.calls) == ["1", "2", "3"]
    assert sorted(state.processed_company_ids) == ["1", "2", "3"]
    assert state.summary["companies_scanned"] == 3
    assert not state.failures

    reloaded = state_store.load(state.run_id)
    assert sorted(reloaded.processed_company_ids) == ["1", "2", "3"]


def test_collect_targets_isolates_failing_company_without_aborting_others(tmp_path: Path):
    companies = [
        CompanyRecord(company_id="1", company_name="A", overdue_days_raw="1"),
        CompanyRecord(company_id="2", company_name="B", overdue_days_raw="1"),
    ]
    results_by_company = {"1": [_search_item("app-1", "1")]}
    client = FakeClient(results_by_company, failing_company_ids=frozenset({"2"}))
    app, state_store = _make_app(tmp_path, client)
    state = state_store.create("run", {}, companies)

    targets = app._collect_targets(state, companies)

    assert [item.app_id for item in targets] == ["app-1"]
    assert state.failures == {"2": "boom-2"}
    assert sorted(state.processed_company_ids) == ["1", "2"]
    assert state.summary["companies_skipped"] == 1


def test_collect_targets_skips_already_processed_companies(tmp_path: Path):
    companies = [
        CompanyRecord(company_id="1", company_name="A", overdue_days_raw="1"),
        CompanyRecord(company_id="2", company_name="B", overdue_days_raw="1"),
    ]
    client = FakeClient({"2": [_search_item("app-2", "2")]})
    app, state_store = _make_app(tmp_path, client)
    state = state_store.create("run", {}, companies)
    state.processed_company_ids.append("1")

    targets = app._collect_targets(state, companies)

    assert [item.app_id for item in targets] == ["app-2"]
    assert client.calls == ["2"]


def test_run_skips_as_duplicate_when_todays_sheet_already_exists(tmp_path: Path):
    companies = [CompanyRecord(company_id="1", company_name="A", overdue_days_raw="1")]
    client = FakeClient({"1": [_search_item("app-1", "1")]})
    app, _ = _make_app(tmp_path, client)

    today_name = build_sheet_name(datetime.now(ZoneInfo(app.settings.excel.timezone)))
    workbook = Workbook()
    workbook.active.title = today_name
    existing_output = tmp_path / "existing_output.xlsm"
    workbook.save(existing_output)
    app.storage.upload_file(existing_output, app.settings.storage.onedrive.output_path)

    result = app.run()

    assert result == {"run_id": None, "sheet_name": None, "skipped": "already_run_today"}
    assert client.calls == []


def test_run_with_clear_cache_ignores_existing_todays_sheet(tmp_path: Path):
    companies = [CompanyRecord(company_id="1", company_name="A", overdue_days_raw="1")]
    client = FakeClient({"1": [_search_item("app-1", "1")]})
    app, _ = _make_app(tmp_path, client)

    today_name = build_sheet_name(datetime.now(ZoneInfo(app.settings.excel.timezone)))
    workbook = Workbook()
    workbook.active.title = today_name
    existing_output = tmp_path / "existing_output.xlsm"
    workbook.save(existing_output)
    app.storage.upload_file(existing_output, app.settings.storage.onedrive.output_path)

    excel_settings = app.settings.excel
    input_workbook = Workbook()
    input_sheet = input_workbook.active
    input_sheet.title = excel_settings.input_sheet_name
    input_sheet.append(
        [excel_settings.company_id_column, excel_settings.company_name_column, excel_settings.overdue_days_column]
    )
    input_sheet.append(["1", "A", "5"])
    existing_input = tmp_path / "existing_input.xlsx"
    input_workbook.save(existing_input)
    app.storage.upload_file(existing_input, app.settings.storage.onedrive.input_path)

    result = app.run(clear_cache=True)

    assert result.get("skipped") is None
    assert client.calls == ["000000001"]
