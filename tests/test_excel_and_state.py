from pathlib import Path

from openpyxl import Workbook, load_workbook

from tender_tracker.config import load_settings
from tender_tracker.excel import read_debtor_companies, write_output_workbook
from tender_tracker.models import PaymentRecord
from tender_tracker.state import RunStateStore
from tender_tracker.storage import LocalStorage


def create_input_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "კომპანიის_საიდენტიფიკაციო_კოდი",
            "კომპანია",
            "ვადაგადაცილებული დღეების რაოდენობა",
        ]
    )
    sheet.append([123456789, "Alpha", 3])
    sheet.append([123456789, "Alpha", 5])
    sheet.append([987654321, "Beta", "-2"])
    sheet.append([111111111, "Gamma", 7])
    workbook.save(path)
    workbook.close()


def test_read_debtor_companies_filters_and_deduplicates(tmp_path: Path):
    workbook_path = tmp_path / "input.xlsx"
    create_input_workbook(workbook_path)
    settings = load_settings("config/settings.yaml")
    companies, names = read_debtor_companies(workbook_path, settings.excel)
    assert [company.company_id for company in companies] == ["123456789", "111111111"]
    assert names == ["Alpha", "Gamma"]


def test_write_output_workbook_creates_expected_columns(tmp_path: Path):
    output = tmp_path / "output.xlsx"
    records = [
        PaymentRecord(
            company_id="123456789",
            company_name="Alpha",
            app_id="678938",
            tender_registration_number="B2B260000023",
            raw_amount="7`500.00 ლარი",
            cleaned_amount=7500.0,
            raw_payment_date="12.03.2026",
            parsed_payment_date=None,
            payment_exists=True,
            warnings=[],
        )
    ]
    sheet_name = write_output_workbook(output, records, ["Alpha"], __import__("datetime").datetime(2026, 3, 31))
    workbook = load_workbook(output, keep_vba=True)
    sheet = workbook[sheet_name]
    assert sheet["A1"].value == "კომპანია"
    assert sheet["B1"].value == "თანხა_რიცხვი"
    assert sheet["C1"].value == "თარიღი_თარიღად"
    assert sheet["E1"].value == "ყველა კომპანია"
    assert sheet["A2"].value == "Alpha"
    assert sheet["E2"].value == "Alpha"
    assert sheet["G2"].value == "7`500.00 ლარი"
    assert sheet["H2"].value == "12.03.2026"
    assert 'G2," ლარი"' in sheet["B2"].value
    assert 'DATE(MID(H2,7,4),MID(H2,4,2),LEFT(H2,2))' in sheet["C2"].value
    assert sheet.column_dimensions["G"].hidden is True
    assert sheet.column_dimensions["H"].hidden is True
    assert len(sheet.tables) == 2
    workbook.close()


def test_write_output_workbook_inserts_new_sheet_first(tmp_path: Path):
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.active.title = "Existing"
    workbook.create_sheet(title="Archive")
    workbook.save(output)
    workbook.close()

    sheet_name = write_output_workbook(output, [], ["Alpha"], __import__("datetime").datetime(2026, 4, 1))

    workbook = load_workbook(output, keep_vba=True)
    assert workbook.sheetnames[0] == sheet_name
    assert workbook.sheetnames[1:] == ["Existing", "Archive"]
    workbook.close()


def test_run_state_store_round_trip(tmp_path: Path):
    root = tmp_path / "storage"
    storage = LocalStorage(root)
    settings = load_settings("config/settings.yaml")
    settings.storage.backend = "local"
    settings.storage.local_root = str(root)
    store = RunStateStore(settings, storage)
    state = store.create("run", {"mode": "run"}, [])
    loaded = store.load(state.run_id)
    assert loaded.run_id == state.run_id
