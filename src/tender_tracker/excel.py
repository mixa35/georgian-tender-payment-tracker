from __future__ import annotations

import calendar
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from tender_tracker.config import ExcelSettings
from tender_tracker.models import CompanyRecord, PaymentRecord

NO_RECORDS_TEXT = "ჩანაწერები არ არის"


class InputWorkbookError(ValueError):
    """Raised when the configured input workbook cannot be read safely."""


def _normalize_company_id(value: object) -> str:
    digits = "".join(ch for ch in str(value).strip() if ch.isdigit())
    return digits.zfill(9) if digits else ""


def _normalize_company_name(value: object) -> str:
    return " ".join(str(value or "").split())


def _is_positive_overdue(value: object) -> bool:
    text = str(value or "").strip()
    if not text or text.startswith("-"):
        return False
    normalized = text.replace(",", ".")
    try:
        return Decimal(normalized) > 0
    except InvalidOperation:
        return False


def _column_index(header: list[str], excel_settings: ExcelSettings) -> dict[str, int]:
    index = {name: idx for idx, name in enumerate(header)}
    required = [
        excel_settings.company_id_column,
        excel_settings.company_name_column,
        excel_settings.overdue_days_column,
    ]
    missing = [name for name in required if name not in index]
    if missing:
        available = [name for name in header if name]
        raise InputWorkbookError(
            "Input workbook is missing required columns: "
            + ", ".join(missing)
            + ". Found columns: "
            + (", ".join(available) if available else "(none)")
        )
    return index


def read_debtor_companies(workbook_path: Path, excel_settings: ExcelSettings) -> tuple[list[CompanyRecord], list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[excel_settings.input_sheet_name] if excel_settings.input_sheet_name else workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []

        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        index = _column_index(header, excel_settings)
        company_id_index = index[excel_settings.company_id_column]
        company_name_index = index[excel_settings.company_name_column]
        overdue_index = index[excel_settings.overdue_days_column]

        unique_by_company: dict[str, CompanyRecord] = {}
        unique_company_names: dict[str, str] = {}
        for row in rows[1:]:
            company_id = _normalize_company_id(row[company_id_index])
            company_name = _normalize_company_name(row[company_name_index])
            overdue_raw = str(row[overdue_index] or "").strip()
            if not company_id or not company_name or not _is_positive_overdue(overdue_raw):
                continue
            unique_by_company.setdefault(
                company_id,
                CompanyRecord(company_id=company_id, company_name=company_name, overdue_days_raw=overdue_raw),
            )
            unique_company_names.setdefault(company_name.casefold(), company_name)

        companies = sorted(unique_by_company.values(), key=lambda item: item.company_name.casefold())
        names = sorted(unique_company_names.values(), key=str.casefold)
        return companies, names
    finally:
        workbook.close()


def build_sheet_name(now: datetime) -> str:
    return f"{now.day} {calendar.month_abbr[now.month]}"


def build_unique_sheet_name(existing_names: list[str], now: datetime) -> str:
    base_name = build_sheet_name(now)
    if base_name not in existing_names:
        return base_name

    suffix = 2
    while True:
        candidate = f"{base_name} ({suffix})"
        if candidate not in existing_names:
            return candidate
        suffix += 1


def write_output_workbook(
    workbook_path: Path,
    records: list[PaymentRecord],
    unique_company_names: list[str],
    run_started_at: datetime,
) -> str:
    workbook = load_workbook(workbook_path, keep_vba=True) if workbook_path.exists() else Workbook()

    sheet_name = build_unique_sheet_name(workbook.sheetnames, run_started_at)
    sheet = workbook.create_sheet(title=sheet_name, index=0)

    headers = {
        "A1": "კომპანია",
        "B1": "თანხა",
        "C1": "თარიღი",
        "D1": "ბმული",
        "F1": "ყველა კომპანია",
    }
    for cell, value in headers.items():
        sheet[cell] = value

    for row_index, record in enumerate(records, start=2):
        sheet[f"A{row_index}"] = record.company_name

        amount_cell = sheet[f"B{row_index}"]
        if record.cleaned_amount is None and record.raw_amount == NO_RECORDS_TEXT:
            amount_cell.value = NO_RECORDS_TEXT
        else:
            amount_cell.value = record.cleaned_amount
            amount_cell.number_format = "0.00"

        date_cell = sheet[f"C{row_index}"]
        if record.parsed_payment_date is not None:
            date_cell.value = record.parsed_payment_date
            date_cell.number_format = "DD/MM/YYYY"

        if record.tender_url:
            link_cell = sheet[f"D{row_index}"]
            link_cell.value = record.tender_url
            link_cell.hyperlink = record.tender_url
            link_cell.style = "Hyperlink"

    for name_row, company_name in enumerate(unique_company_names, start=2):
        sheet[f"F{name_row}"] = company_name

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 45
    sheet.column_dimensions["E"].width = 3
    sheet.column_dimensions["F"].width = 28
    sheet.column_dimensions["G"].width = 3

    if records:
        tender_table = Table(
            displayName=f"TenderData_{run_started_at:%Y%m%d}_{len(workbook.sheetnames)}",
            ref=f"A1:D{len(records) + 1}",
        )
        tender_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(tender_table)

    if unique_company_names:
        company_table = Table(
            displayName=f"DebtorCompanies_{run_started_at:%Y%m%d}_{len(workbook.sheetnames)}",
            ref=f"F1:F{len(unique_company_names) + 1}",
        )
        company_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium3",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(company_table)

    workbook.save(workbook_path)
    workbook.close()
    return sheet_name
